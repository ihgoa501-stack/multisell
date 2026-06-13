"""商品管理 - 路由"""

from fastapi import APIRouter, Depends, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from datetime import datetime
import io, openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from sqlalchemy.ext.asyncio import AsyncSession
from app.database import get_db
from app.common import Result, PageResult, ProductStatus
from app.core.schemas import ProductCreate, ProductUpdate, ProductQuery, ProductVO
from app.core.service import ProductService

router = APIRouter(tags=["商品管理"])


def product_to_vo(product) -> ProductVO:
    """商品模型转VO"""
    status_name = ProductStatus.STATUS_MAP.get(product.status, "未知")
    category_name = product.category.name if product.category else None
    # 获取品牌名称（如果有brand_id）
    brand_name = None
    if hasattr(product, 'brand_id') and product.brand_id:
        # 这里不额外查询，前端会通过/brands/all获取
        pass
    return ProductVO(
        id=product.id,
        name=product.name,
        subtitle=product.subtitle,
        description=product.description,
        brand_id=product.brand_id,
        category_id=product.category_id,
        category_name=category_name,
        brand_name=None,
        unit=product.unit,
        status=product.status,
        status_name=status_name,
        main_image=product.main_image,
        images=product.images,
        ai_status=product.ai_status,
        platform_statuses=product.platform_statuses,
        created_at=product.created_at,
        updated_at=product.updated_at,
    )


@router.post("/products/batch/status", summary="批量修改商品状态")
async def batch_update_status(data: dict, db: AsyncSession = Depends(get_db)):
    """批量修改商品状态"""
    from sqlalchemy import update
    from app.models import Product

    ids = data.get("ids", [])
    status = data.get("status", 0)
    if not ids:
        return Result.bad_request("请选择商品")

    stmt = update(Product).where(Product.id.in_(ids)).values(status=status)
    await db.execute(stmt)
    await db.flush()
    return Result.ok({"affected": len(ids)})


@router.post("/products/batch/delete", summary="批量删除商品")
async def batch_delete_products(data: dict, db: AsyncSession = Depends(get_db)):
    """批量删除商品"""
    from sqlalchemy import delete
    from app.models import Product

    ids = data.get("ids", [])
    if not ids:
        return Result.bad_request("请选择商品")

    stmt = delete(Product).where(Product.id.in_(ids))
    result = await db.execute(stmt)
    await db.flush()
    return Result.ok({"affected": result.rowcount})


@router.post("/products", summary="创建商品")
async def create_product(data: ProductCreate, db: AsyncSession = Depends(get_db)):
    product = await ProductService.create(db, data)
    return Result.ok(product_to_vo(product))


@router.get("/products/export", summary="导出商品Excel")
async def export_products(
    name: str = Query(None, description="按名称筛选"),
    category_id: int = Query(None, description="按分类筛选"),
    status: int = Query(None, description="按状态筛选"),
    db: AsyncSession = Depends(get_db),
):
    """导出商品列表为 Excel 文件"""
    from app.core.service import ProductService

    # 分批拉取所有商品
    all_products = []
    page = 1
    while True:
        q = ProductQuery(name=name, category_id=category_id, status=status, page=page, page_size=100)
        products, total = await ProductService.list_products(db, q)
        all_products.extend(products)
        if len(all_products) >= total:
            break
        page += 1

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "商品列表"

    headers = ["ID", "商品名称", "副标题", "分类", "单位", "状态", "创建时间"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    status_map = {0: "草稿", 1: "上架", 2: "下架"}
    for row_idx, p in enumerate(products, 2):
        category_name = p.category.name if p.category else ""
        ws.cell(row=row_idx, column=1, value=p.id)
        ws.cell(row=row_idx, column=2, value=p.name)
        ws.cell(row=row_idx, column=3, value=p.subtitle or "")
        ws.cell(row=row_idx, column=4, value=category_name)
        ws.cell(row=row_idx, column=5, value=p.unit)
        ws.cell(row=row_idx, column=6, value=status_map.get(p.status, "未知"))
        ws.cell(row=row_idx, column=7, value=p.created_at.strftime("%Y-%m-%d %H:%M") if p.created_at else "")
        for col in range(1, 8):
            ws.cell(row=row_idx, column=col).border = thin_border
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 8
    ws.column_dimensions["G"].width = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=products_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"},
    )


@router.post("/products/import", summary="从Excel导入商品")
async def import_products(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
):
    """从 Excel 文件导入商品"""
    if not file.filename or not file.filename.endswith(('.xlsx', '.xls')):
        return Result.bad_request("请上传 .xlsx 或 .xls 文件")
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    if ws.max_row < 2:
        return Result.bad_request("Excel 文件为空")
    imported = 0
    errors = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[1] if len(row) > 1 else None
        if not name:
            continue
        try:
            product_data = ProductCreate(
                name=str(name),
                subtitle=str(row[2]) if len(row) > 2 and row[2] else None,
                unit=str(row[4]) if len(row) > 4 and row[4] else "件",
                status=1 if len(row) > 5 and row[5] == "上架" else 0 if row[5] == "草稿" else 2,
            )
            await ProductService.create(db, product_data)
            imported += 1
        except Exception as e:
            errors.append(f"第{row[0] if row[0] else row_idx}行: {str(e)}")
    return Result.ok({"imported": imported, "errors": errors, "total": ws.max_row - 1})


@router.put("/products/{product_id}", summary="更新商品")
async def update_product(product_id: int, data: ProductUpdate, db: AsyncSession = Depends(get_db)):
    product = await ProductService.update(db, product_id, data)
    if not product:
        return Result.not_found("商品不存在")
    return Result.ok(product_to_vo(product))


@router.get("/products/{product_id}/detail", summary="商品聚合详情")
async def get_product_detail(product_id: int, db: AsyncSession = Depends(get_db)):
    """聚合返回商品详情 + SKU列表 + 价格 + 库存 + 供应商"""
    from sqlalchemy import select
    from app.models import Sku, Price, Inventory, Brand, ProductSupplier, Supplier, ProductListing, Platform

    product = await ProductService.get_by_id(db, product_id)
    if not product:
        return Result.not_found("商品不存在")

    vo = product_to_vo(product)

    # 品牌名称
    if product.brand_id:
        brand = await db.get(Brand, product.brand_id)
        if brand:
            vo.brand_name = brand.name

    # SKU列表
    stmt = select(Sku).where(Sku.product_id == product_id)
    skus_result = await db.execute(stmt)
    skus = skus_result.scalars().all()
    sku_list = []
    for sku in skus:
        s = {
            "id": sku.id,
            "code": sku.code,
            "barcode": sku.barcode,
            "spec_desc": sku.spec_desc,
            "price": float(sku.price) if sku.price else None,
            "market_price": float(sku.market_price) if sku.market_price else None,
            "stock": sku.stock or 0,
            "image": sku.image,
            "status": sku.status,
        }
        # 获取每个SKU的当前售价
        price_stmt = select(Price).where(
            Price.sku_id == sku.id,
            Price.price_type == "sale_price",
            Price.status == 1,
        ).order_by(Price.created_at.desc()).limit(1)
        price_res = await db.execute(price_stmt)
        current_price = price_res.scalar_one_or_none()
        if current_price:
            s["sale_price"] = float(current_price.price)
        sku_list.append(s)

    # 库存信息
    inv_list = []
    sku_ids = [s.id for s in skus]
    if sku_ids:
        inv_stmt = select(Inventory).where(Inventory.sku_id.in_(sku_ids))
        inv_res = await db.execute(inv_stmt)
        for inv in inv_res.scalars().all():
            inv_list.append({
                "id": inv.id,
                "sku_id": inv.sku_id,
                "warehouse": inv.warehouse,
                "quantity": inv.quantity,
                "safety_stock": inv.safety_stock,
            })

    # 供应商列表
    ps_stmt = select(ProductSupplier, Supplier.name).join(
        Supplier, ProductSupplier.supplier_id == Supplier.id
    ).where(ProductSupplier.product_id == product_id)
    ps_res = await db.execute(ps_stmt)
    suppliers = []
    for ps, name in ps_res.all():
        suppliers.append({
            "id": ps.id,
            "supplier_id": ps.supplier_id,
            "supplier_name": name,
            "supply_price": float(ps.supply_price) if ps.supply_price else None,
        })

    # 发布状态
    listing_stmt = (
        select(ProductListing, Platform.name, Platform.code)
        .join(Platform, ProductListing.platform_id == Platform.id)
        .where(ProductListing.product_id == product_id)
        .order_by(Platform.sort_order)
    )
    listing_res = await db.execute(listing_stmt)
    listings = []
    for listing, plat_name, plat_code in listing_res.all():
        listings.append({
            "id": listing.id,
            "platform_id": listing.platform_id,
            "platform_name": plat_name,
            "platform_code": plat_code,
            "platform_product_id": listing.platform_product_id,
            "status": listing.status,
            "platform_url": listing.platform_url,
            "last_sync_at": listing.last_sync_at.isoformat() if listing.last_sync_at else None,
        })

    return Result.ok({
        "product": vo.model_dump(),
        "skus": sku_list,
        "inventory": inv_list,
        "suppliers": suppliers,
        "listings": listings,
    })


@router.post("/products/{product_id}/duplicate", summary="复制商品")
async def duplicate_product(product_id: int, db: AsyncSession = Depends(get_db)):
    """基于已有商品复制一个新商品（复制基本信息，不复制SKU/价格/库存）"""
    from app.models import Product
    
    product = await ProductService.get_by_id(db, product_id)
    if not product:
        return Result.not_found("商品不存在")

    new_product = Product(
        name=f"{product.name} (副本)",
        subtitle=product.subtitle,
        description=product.description,
        brand_id=product.brand_id,
        category_id=product.category_id,
        unit=product.unit,
        main_image=product.main_image,
        images=product.images,
        status=0,  # 新商品默认为草稿
    )
    db.add(new_product)
    await db.flush()
    await db.refresh(new_product)

    return Result.ok(product_to_vo(new_product))


@router.get("/products/{product_id}", summary="商品详情")
async def get_product(product_id: int, db: AsyncSession = Depends(get_db)):
    product = await ProductService.get_by_id(db, product_id)
    if not product:
        return Result.not_found("商品不存在")
    return Result.ok(product_to_vo(product))


@router.get("/products", summary="商品列表")
async def list_products(
    name: str = Query(None, description="商品名称"),
    category_id: int = Query(None, description="分类ID"),
    status: int = Query(None, description="状态"),
    brand_id: int = Query(None, description="品牌ID"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
):
    query = ProductQuery(name=name, category_id=category_id, status=status, brand_id=brand_id, page=page, page_size=page_size)
    products, total = await ProductService.list_products(db, query)
    items = [product_to_vo(p) for p in products]
    return PageResult.ok(items, total, page, page_size)


@router.delete("/products/{product_id}", summary="删除商品")
async def delete_product(product_id: int, db: AsyncSession = Depends(get_db)):
    ok = await ProductService.delete(db, product_id)
    if not ok:
        return Result.not_found("商品不存在")
    return Result.ok(message="删除成功")


@router.post("/products/{product_id}/ai-enhance", summary="AI优化商品信息")
async def ai_enhance_product(product_id: int, db: AsyncSession = Depends(get_db)):
    """AI生成优化标题、描述和SEO关键词（当前为模拟数据，后续可接入LLM）"""
    from app.models import Product
    from datetime import datetime

    product = await db.get(Product, product_id)
    if not product:
        return Result.not_found("商品不存在")

    # 模拟AI生成结果（基于商品名称构建）
    name = product.name or ""
    
    # 生成优化标题（模拟）
    enhanced_title = f"{name} - 高品质正品保障 厂家直销批发"

    # 生成优化描述（模拟）
    enhanced_desc = (
        f"【{name}】品质保障，正品货源。\n\n"
        f"产品特色：\n"
        f"✅ 优质材料，经久耐用\n"
        f"✅ 严格品控，质量可靠\n"
        f"✅ 厂家直供，价格优惠\n"
        f"✅ 支持批发/零售，快速发货\n\n"
        f"欢迎联系我们获取更多产品信息！"
    )

    # 生成SEO关键词（模拟）
    keywords = [
        name,
        f"{name} 批发",
        f"{name} 厂家",
        f"{name} 价格",
    ]

    # 保存到商品记录
    product.ai_title = enhanced_title
    product.ai_description = enhanced_desc
    product.seo_keywords = keywords
    product.ai_status = "completed"
    await db.flush()

    return Result.ok({
        "enhanced_title": enhanced_title,
        "enhanced_description": enhanced_desc,
        "seo_keywords": keywords,
        "ai_status": "completed",
        "message": "AI优化完成，请检查并确认",
    })


