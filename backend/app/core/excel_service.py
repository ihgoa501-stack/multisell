"""Excel 导出导入 — 服务层"""

import io
from typing import Optional

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.schemas import ProductCreate, ProductQuery
from app.core.service import ProductService


def _to_positive_float(value, field_name: str, row_idx: int) -> Optional[float]:
    """将单元格值转为正浮点数，无效时报行级错误"""
    if value in (None, ""):
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        raise ValueError(f"第{row_idx}行: {field_name}必须是数字")
    if number <= 0:
        raise ValueError(f"第{row_idx}行: {field_name}必须大于0")
    return number


def _cell_value(row, header_map: dict, name: str):
    """按表头名称读取单元格值"""
    index = header_map.get(name)
    if index is None or index >= len(row):
        return None
    return row[index]


class ExcelService:
    """商品 Excel 导出导入"""

    STYLE_HEADER_FONT = Font(bold=True, color="FFFFFF")
    STYLE_HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    STYLE_THIN_BORDER = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    EXPORT_HEADERS = [
        "ID", "商品名称", "副标题", "分类", "单位", "状态",
        "商品长(cm)", "商品宽(cm)", "商品高(cm)", "商品重量(kg)",
        "包装长(cm)", "包装宽(cm)", "包装高(cm)", "包装重量(kg)",
        "货品类型", "物流状态", "创建时间",
    ]

    TEMPLATE_HEADERS = [
        "商品名称", "副标题", "单位", "状态",
        "商品长(cm)", "商品宽(cm)", "商品高(cm)", "商品重量(kg)",
        "包装长(cm)", "包装宽(cm)", "包装高(cm)", "包装重量(kg)",
        "货品类型",
    ]

    @staticmethod
    def _apply_header_style(ws, headers: list[str]):
        """给表头行应用样式"""
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = ExcelService.STYLE_HEADER_FONT
            cell.fill = ExcelService.STYLE_HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = ExcelService.STYLE_THIN_BORDER

    @staticmethod
    async def export_products(
        db: AsyncSession,
        name: Optional[str] = None,
        category_id: Optional[int] = None,
        status: Optional[int] = None,
        brand_id: Optional[int] = None,
        cargo_type: Optional[str] = None,
        logistics_status: Optional[str] = None,
    ) -> io.BytesIO:
        """导出商品列表为 Excel 文件，返回 BytesIO"""
        # 分批拉取所有商品
        all_products = []
        page = 1
        while True:
            q = ProductQuery(
                name=name, category_id=category_id, status=status, brand_id=brand_id,
                cargo_type=cargo_type, logistics_status=logistics_status,
                page=page, page_size=100,
            )
            products, total = await ProductService.list_products(db, q)
            all_products.extend(products)
            if len(all_products) >= total:
                break
            page += 1

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "商品列表"

        headers = ExcelService.EXPORT_HEADERS
        ExcelService._apply_header_style(ws, headers)

        status_map = {0: "草稿", 1: "上架", 2: "下架"}
        cargo_type_map = {"normal": "普通", "battery": "带电", "liquid": "液体", "sensitive": "敏感"}
        for row_idx, p in enumerate(all_products, 2):
            category_name = p.category.name if p.category else ""
            logistics_complete = (
                p.package_length_cm and p.package_width_cm and p.package_height_cm and p.package_weight_kg
                and float(p.package_length_cm) > 0 and float(p.package_width_cm) > 0
                and float(p.package_height_cm) > 0 and float(p.package_weight_kg) > 0
            )
            logistics_label = "物流完整" if logistics_complete else "物流不完整"

            vals = [
                p.id, p.name, p.subtitle or "", category_name, p.unit,
                status_map.get(p.status, "未知"),
                _to_positive_float(p.product_length_cm, "", 0) if p.product_length_cm else None,
                _to_positive_float(p.product_width_cm, "", 0) if p.product_width_cm else None,
                _to_positive_float(p.product_height_cm, "", 0) if p.product_height_cm else None,
                _to_positive_float(p.product_weight_kg, "", 0) if p.product_weight_kg else None,
                _to_positive_float(p.package_length_cm, "", 0) if p.package_length_cm else None,
                _to_positive_float(p.package_width_cm, "", 0) if p.package_width_cm else None,
                _to_positive_float(p.package_height_cm, "", 0) if p.package_height_cm else None,
                _to_positive_float(p.package_weight_kg, "", 0) if p.package_weight_kg else None,
                cargo_type_map.get(p.cargo_type, p.cargo_type or ""),
                logistics_label,
                p.created_at.strftime("%Y-%m-%d %H:%M") if p.created_at else "",
            ]
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.border = ExcelService.STYLE_THIN_BORDER

        col_widths = {
            "A": 8, "B": 40, "C": 30, "D": 15, "E": 8, "F": 8,
            "G": 12, "H": 10, "I": 10, "J": 12,
            "K": 12, "L": 10, "M": 10, "N": 12,
            "O": 12, "P": 12, "Q": 20,
        }
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def export_template() -> io.BytesIO:
        """下载商品导入模板（空表头），返回 BytesIO"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "商品导入模板"

        headers = ExcelService.TEMPLATE_HEADERS
        ExcelService._apply_header_style(ws, headers)

        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 8
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 10
        ws.column_dimensions["G"].width = 10
        ws.column_dimensions["H"].width = 12
        ws.column_dimensions["I"].width = 12
        ws.column_dimensions["J"].width = 10
        ws.column_dimensions["K"].width = 10
        ws.column_dimensions["L"].width = 12
        ws.column_dimensions["M"].width = 12

        # 数据验证：状态列下拉
        from openpyxl.worksheet.datavalidation import DataValidation
        dv_status = DataValidation(type="list", formula1='"草稿,上架,下架"', allow_blank=True)
        dv_status.error = "请选择: 草稿 / 上架 / 下架"
        dv_status.errorTitle = "状态错误"
        ws.add_data_validation(dv_status)
        dv_status.add("D2:D1048576")

        # 数据验证：货品类型下拉
        dv_cargo = DataValidation(type="list", formula1='"normal,battery,liquid,sensitive"', allow_blank=True)
        dv_cargo.error = "请选择: normal / battery / liquid / sensitive"
        dv_cargo.errorTitle = "货品类型错误"
        ws.add_data_validation(dv_cargo)
        dv_cargo.add("M2:M1048576")

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    async def import_products(
        db: AsyncSession,
        file_bytes: bytes,
    ) -> dict:
        """从 Excel 文件内容导入商品，返回 {imported, errors, total}"""
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
        ws = wb.active
        if ws.max_row < 2:
            return {"imported": 0, "errors": ["Excel 文件为空"], "total": 0}

        # 构建表头映射
        headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
        header_map = {name: idx for idx, name in enumerate(headers)}

        imported = 0
        errors = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            name = _cell_value(row, header_map, "商品名称")
            if not name:
                continue
            try:
                # 读取状态
                raw_status = _cell_value(row, header_map, "状态")
                if raw_status == "上架":
                    status_val = 1
                elif raw_status == "草稿":
                    status_val = 0
                elif raw_status == "下架":
                    status_val = 2
                else:
                    status_val = 0
                # 读取单位
                raw_unit = _cell_value(row, header_map, "单位")
                unit_val = str(raw_unit).strip() if raw_unit else "件"
                # 读取货品类型
                raw_cargo = _cell_value(row, header_map, "货品类型")
                cargo_val = str(raw_cargo).strip() if raw_cargo else "normal"

                product_data = ProductCreate(
                    name=str(name).strip(),
                    subtitle=str(_cell_value(row, header_map, "副标题")).strip()
                    if _cell_value(row, header_map, "副标题") else None,
                    unit=unit_val,
                    status=status_val,
                    product_length_cm=_to_positive_float(
                        _cell_value(row, header_map, "商品长(cm)"), "商品长(cm)", row_idx,
                    ),
                    product_width_cm=_to_positive_float(
                        _cell_value(row, header_map, "商品宽(cm)"), "商品宽(cm)", row_idx,
                    ),
                    product_height_cm=_to_positive_float(
                        _cell_value(row, header_map, "商品高(cm)"), "商品高(cm)", row_idx,
                    ),
                    product_weight_kg=_to_positive_float(
                        _cell_value(row, header_map, "商品重量(kg)"), "商品重量(kg)", row_idx,
                    ),
                    package_length_cm=_to_positive_float(
                        _cell_value(row, header_map, "包装长(cm)"), "包装长(cm)", row_idx,
                    ),
                    package_width_cm=_to_positive_float(
                        _cell_value(row, header_map, "包装宽(cm)"), "包装宽(cm)", row_idx,
                    ),
                    package_height_cm=_to_positive_float(
                        _cell_value(row, header_map, "包装高(cm)"), "包装高(cm)", row_idx,
                    ),
                    package_weight_kg=_to_positive_float(
                        _cell_value(row, header_map, "包装重量(kg)"), "包装重量(kg)", row_idx,
                    ),
                    cargo_type=cargo_val,
                )
                await ProductService.create(db, product_data)
                imported += 1
            except Exception as e:
                errors.append(f"第{row_idx}行: {str(e)}")

        return {"imported": imported, "errors": errors, "total": ws.max_row - 1}
