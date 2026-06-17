import io
import json
from datetime import datetime
from typing import Optional

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import ImportBatch, ImportBatchRow, Product, Sku, Price, PriceChangeLog, Inventory, InventoryLog

IMPORT_TYPES = frozenset({"product", "sku", "price", "inventory"})

TEMPLATE_HEADERS = {
    "product": [
        "商品名称", "副标题", "单位", "状态",
        "商品长(cm)", "商品宽(cm)", "商品高(cm)", "商品重量(kg)",
        "包装长(cm)", "包装宽(cm)", "包装高(cm)", "包装重量(kg)",
        "货品类型",
    ],
    "sku": [
        "SKU编码", "所属商品ID", "条码", "规格描述", "规格值(JSON)",
        "销售价", "成本价", "市场价", "库存", "状态", "重量(kg)",
    ],
    "price": [
        "SKU编码", "价格类型", "价格", "生效时间", "失效时间",
    ],
    "inventory": [
        "SKU编码", "仓库", "货位", "数量", "模式", "安全库存",
    ],
}

STATUS_MAP = {"草稿": 0, "上架": 1, "下架": 2}
VALID_PRICE_TYPES = {"sale_price", "market_price", "cost_price", "vip_price", "wholesale_price"}


class ImportStyle:
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    THIN_BORDER = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )


class ImportBatchService:

    @staticmethod
    def _apply_header_style(ws, headers: list[str]):
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = ImportStyle.HEADER_FONT
            cell.fill = ImportStyle.HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = ImportStyle.THIN_BORDER

    @staticmethod
    def generate_template(import_type: str) -> io.BytesIO:
        """Generate empty xlsx template for given import type."""
        headers = TEMPLATE_HEADERS[import_type]
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{import_type}导入模板"
        ImportBatchService._apply_header_style(ws, headers)
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def _cell_value(row, header_map: dict, name: str):
        index = header_map.get(name)
        if index is None or index >= len(row):
            return None
        return row[index]

    @staticmethod
    def _to_float(value) -> Optional[float]:
        if value is None or value == "":
            return None
        try:
            return float(value)
        except (TypeError, ValueError):
            return None

    @staticmethod
    def _to_int(value) -> Optional[int]:
        if value is None or value == "":
            return None
        try:
            return int(value)
        except (TypeError, ValueError):
            return None

    @staticmethod
    def _validate_row(import_type: str, row, header_map: dict, row_idx: int) -> Optional[str]:
        errors = []

        if import_type == "product":
            name = ImportBatchService._cell_value(row, header_map, "商品名称")
            if not name or not str(name).strip():
                errors.append("商品名称不能为空")

        elif import_type == "sku":
            code = ImportBatchService._cell_value(row, header_map, "SKU编码")
            product_id = ImportBatchService._cell_value(row, header_map, "所属商品ID")
            if not code and not product_id:
                errors.append("SKU编码或所属商品ID至少填写一个")

        elif import_type == "price":
            sku_code = ImportBatchService._cell_value(row, header_map, "SKU编码")
            if not sku_code:
                errors.append("SKU编码不能为空")
            price_type = ImportBatchService._cell_value(row, header_map, "价格类型")
            if price_type not in VALID_PRICE_TYPES:
                errors.append(f"价格类型无效(可选: {', '.join(sorted(VALID_PRICE_TYPES))})")
            price_val = ImportBatchService._cell_value(row, header_map, "价格")
            if price_val is None or price_val == "":
                errors.append("价格不能为空")
            else:
                try:
                    if float(price_val) <= 0:
                        errors.append("价格必须大于0")
                except (TypeError, ValueError):
                    errors.append("价格必须是数字")

        elif import_type == "inventory":
            sku_code = ImportBatchService._cell_value(row, header_map, "SKU编码")
            if not sku_code:
                errors.append("SKU编码不能为空")
            qty = ImportBatchService._cell_value(row, header_map, "数量")
            if qty is None or qty == "":
                errors.append("数量不能为空")
            else:
                try:
                    if int(qty) < 0:
                        errors.append("数量不能为负数")
                except (TypeError, ValueError):
                    errors.append("数量必须是整数")
            mode = ImportBatchService._cell_value(row, header_map, "模式")
            if mode and str(mode).strip().lower() not in ("overwrite", "delta"):
                errors.append("模式无效(可选: overwrite/delta)")

        return "; ".join(errors) if errors else None

    @staticmethod
    async def preview(
        db: AsyncSession,
        import_type: str,
        file_bytes: bytes,
        file_name: str,
        operator: str,
    ) -> dict:
        """Parse and validate Excel file, store rows, return preview result."""
        batch = ImportBatch(
            type=import_type,
            file_name=file_name,
            status="pending",
            created_by=operator,
        )
        db.add(batch)
        await db.flush()
        await db.refresh(batch)

        wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
        ws = wb.active
        if ws.max_row < 2:
            batch.status = "failed"
            batch.error_summary = "Excel文件为空"
            await db.flush()
            return {
                "batch_id": batch.id,
                "type": import_type,
                "total_rows": 0,
                "valid_rows": 0,
                "error_rows": 0,
                "errors": [{"row_index": 0, "error_message": "Excel文件为空"}],
            }

        headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
        header_map = {name: idx for idx, name in enumerate(headers)}

        errors = []
        total_rows = 0
        valid_rows = 0
        error_rows = 0

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            total_rows += 1
            raw_data = {}
            for h, idx in header_map.items():
                if idx < len(row):
                    raw_data[h] = row[idx]

            error_msg = ImportBatchService._validate_row(import_type, row, header_map, row_idx)

            row_record = ImportBatchRow(
                batch_id=batch.id,
                row_index=row_idx,
                raw_data=raw_data,
                status="error" if error_msg else "pending",
                error_message=error_msg,
            )
            db.add(row_record)

            if error_msg:
                error_rows += 1
                errors.append({"row_index": row_idx, "error_message": error_msg})
            else:
                valid_rows += 1

        batch.total_rows = total_rows
        batch.status = "previewed"
        batch.error_count = error_rows
        batch.success_count = 0
        await db.flush()

        return {
            "batch_id": batch.id,
            "type": import_type,
            "total_rows": total_rows,
            "valid_rows": valid_rows,
            "error_rows": error_rows,
            "errors": errors,
        }

    @staticmethod
    async def commit(db: AsyncSession, batch_id: int, operator: str) -> dict:
        """Execute import for a previewed batch."""
        batch = await db.get(ImportBatch, batch_id)
        if not batch:
            raise ValueError("批次不存在")
        if batch.status != "previewed":
            raise ValueError(f"批次状态不是previewed，当前: {batch.status}")

        stmt = select(ImportBatchRow).where(
            ImportBatchRow.batch_id == batch_id,
            ImportBatchRow.status == "pending",
        ).order_by(ImportBatchRow.row_index)
        result = await db.execute(stmt)
        pending_rows = list(result.scalars().all())

        success_count = 0
        error_count = 0
        errors = []
        import_type = batch.type

        for row_record in pending_rows:
            try:
                if import_type == "product":
                    await ImportBatchService._commit_product(db, row_record)
                elif import_type == "sku":
                    await ImportBatchService._commit_sku(db, row_record)
                elif import_type == "price":
                    await ImportBatchService._commit_price(db, row_record, operator)
                elif import_type == "inventory":
                    await ImportBatchService._commit_inventory(db, row_record, operator)

                row_record.status = "success"
                success_count += 1
            except Exception as e:
                row_record.status = "error"
                row_record.error_message = str(e)
                error_count += 1
                errors.append({"row_index": row_record.row_index, "error_message": str(e)})

        batch.success_count = success_count
        batch.error_count = error_count
        batch.status = "committed"
        if error_count > 0:
            batch.error_summary = f"成功{success_count}行，失败{error_count}行"
        else:
            batch.error_summary = f"全部成功({success_count}行)"
        await db.flush()

        return {
            "batch_id": batch.id,
            "type": import_type,
            "success_count": success_count,
            "error_count": error_count,
            "imported": success_count,
            "errors": errors,
        }

    @staticmethod
    async def _commit_product(db: AsyncSession, row_record: ImportBatchRow):
        raw = row_record.raw_data or {}
        raw_status = str(raw.get("状态", "")).strip() if raw.get("状态") else ""
        status_val = STATUS_MAP.get(raw_status, 0)

        product = Product(
            name=str(raw.get("商品名称", "")).strip(),
            subtitle=str(raw.get("副标题", "")).strip() if raw.get("副标题") else None,
            unit=str(raw.get("单位", "件")).strip(),
            status=status_val,
            product_length_cm=ImportBatchService._to_float(raw.get("商品长(cm)")),
            product_width_cm=ImportBatchService._to_float(raw.get("商品宽(cm)")),
            product_height_cm=ImportBatchService._to_float(raw.get("商品高(cm)")),
            product_weight_kg=ImportBatchService._to_float(raw.get("商品重量(kg)")),
            package_length_cm=ImportBatchService._to_float(raw.get("包装长(cm)")),
            package_width_cm=ImportBatchService._to_float(raw.get("包装宽(cm)")),
            package_height_cm=ImportBatchService._to_float(raw.get("包装高(cm)")),
            package_weight_kg=ImportBatchService._to_float(raw.get("包装重量(kg)")),
            cargo_type=str(raw.get("货品类型", "normal")).strip(),
        )
        db.add(product)
        await db.flush()

    @staticmethod
    async def _commit_sku(db: AsyncSession, row_record: ImportBatchRow):
        raw = row_record.raw_data or {}
        code = str(raw.get("SKU编码", "")).strip() if raw.get("SKU编码") else None
        product_id_val = ImportBatchService._to_int(raw.get("所属商品ID"))

        sku = None
        if code:
            stmt = select(Sku).where(Sku.code == code)
            result = await db.execute(stmt)
            sku = result.scalar_one_or_none()

        if not sku and product_id_val:
            sku = await db.get(Sku, product_id_val)

        if sku:
            if code:
                sku.code = code
            if raw.get("条码"):
                sku.barcode = str(raw.get("条码")).strip()
            if raw.get("规格描述"):
                sku.spec_desc = str(raw.get("规格描述")).strip()
            if raw.get("规格值(JSON)"):
                try:
                    sku.spec_values = json.loads(str(raw.get("规格值(JSON)")))
                except json.JSONDecodeError:
                    pass
            price_val = ImportBatchService._to_float(raw.get("销售价"))
            if price_val is not None:
                sku.price = price_val
            cost_val = ImportBatchService._to_float(raw.get("成本价"))
            if cost_val is not None:
                sku.cost_price = cost_val
            market_val = ImportBatchService._to_float(raw.get("市场价"))
            if market_val is not None:
                sku.market_price = market_val
            stock_val = ImportBatchService._to_int(raw.get("库存"))
            if stock_val is not None:
                sku.stock = stock_val
            status_val = ImportBatchService._to_int(raw.get("状态"))
            if status_val is not None:
                sku.status = status_val
            weight_val = ImportBatchService._to_float(raw.get("重量(kg)"))
            if weight_val is not None:
                sku.weight = weight_val
        else:
            spec_values = None
            if raw.get("规格值(JSON)"):
                try:
                    spec_values = json.loads(str(raw.get("规格值(JSON)")))
                except json.JSONDecodeError:
                    pass
            sku = Sku(
                product_id=product_id_val or 0,
                code=code,
                barcode=str(raw.get("条码", "")).strip() if raw.get("条码") else None,
                spec_desc=str(raw.get("规格描述", "")).strip() if raw.get("规格描述") else None,
                spec_values=spec_values,
                price=ImportBatchService._to_float(raw.get("销售价")) or 0,
                cost_price=ImportBatchService._to_float(raw.get("成本价")) or 0,
                market_price=ImportBatchService._to_float(raw.get("市场价")) or 0,
                stock=ImportBatchService._to_int(raw.get("库存")) or 0,
                status=ImportBatchService._to_int(raw.get("状态")) or 1,
                weight=ImportBatchService._to_float(raw.get("重量(kg)")) or 0,
            )
            db.add(sku)

        await db.flush()

    @staticmethod
    async def _commit_price(db: AsyncSession, row_record: ImportBatchRow, operator: str):
        raw = row_record.raw_data or {}
        sku_code = str(raw.get("SKU编码", "")).strip()
        price_type = str(raw.get("价格类型", "")).strip()
        price_amount = float(raw.get("价格"))

        stmt = select(Sku).where(Sku.code == sku_code)
        result = await db.execute(stmt)
        sku = result.scalar_one_or_none()
        if not sku:
            raise ValueError(f"SKU编码 '{sku_code}' 不存在")

        stmt = select(Price).where(
            Price.sku_id == sku.id,
            Price.price_type == price_type,
        )
        result = await db.execute(stmt)
        existing = result.scalar_one_or_none()

        old_price = float(existing.price) if existing else None

        if existing:
            existing.price = price_amount
            if raw.get("生效时间"):
                existing.start_time = datetime.fromisoformat(str(raw.get("生效时间")))
            if raw.get("失效时间"):
                existing.end_time = datetime.fromisoformat(str(raw.get("失效时间")))
        else:
            start_time = datetime.fromisoformat(str(raw.get("生效时间"))) if raw.get("生效时间") else None
            end_time = datetime.fromisoformat(str(raw.get("失效时间"))) if raw.get("失效时间") else None
            price_obj = Price(
                sku_id=sku.id,
                price_type=price_type,
                price=price_amount,
                start_time=start_time,
                end_time=end_time,
            )
            db.add(price_obj)

        log = PriceChangeLog(
            sku_id=sku.id,
            old_price=old_price,
            new_price=price_amount,
            price_type=price_type,
            change_type="batch",
            operator=operator,
            remark="批量导入调价",
        )
        db.add(log)
        await db.flush()

    @staticmethod
    async def _commit_inventory(db: AsyncSession, row_record: ImportBatchRow, operator: str):
        raw = row_record.raw_data or {}
        sku_code = str(raw.get("SKU编码", "")).strip()
        warehouse = str(raw.get("仓库", "默认仓库")).strip()
        location = str(raw.get("货位", "")).strip() if raw.get("货位") else None
        quantity = int(raw.get("数量"))
        mode = str(raw.get("模式", "overwrite")).strip().lower()
        safety_stock = ImportBatchService._to_int(raw.get("安全库存"))

        stmt = select(Sku).where(Sku.code == sku_code)
        result = await db.execute(stmt)
        sku = result.scalar_one_or_none()
        if not sku:
            raise ValueError(f"SKU编码 '{sku_code}' 不存在")

        stmt = select(Inventory).where(Inventory.sku_id == sku.id)
        result = await db.execute(stmt)
        inv = result.scalar_one_or_none()

        before_qty = inv.quantity if inv else 0

        if mode == "delta":
            if inv:
                new_qty = (inv.quantity or 0) + quantity
                change_qty = quantity
            else:
                new_qty = quantity
                change_qty = quantity
        else:
            new_qty = quantity
            change_qty = quantity - before_qty

        if inv:
            inv.quantity = new_qty
            inv.warehouse = warehouse
            if location:
                inv.location = location
            if safety_stock is not None:
                inv.safety_stock = safety_stock
        else:
            inv = Inventory(
                sku_id=sku.id,
                warehouse=warehouse,
                location=location,
                quantity=new_qty,
                locked_quantity=0,
                safety_stock=safety_stock or 0,
            )
            db.add(inv)

        await db.flush()

        log = InventoryLog(
            sku_id=sku.id,
            change_type="adjust",
            change_qty=change_qty,
            before_qty=before_qty,
            after_qty=new_qty,
            remark=f"批量导入: {mode}模式",
            operator=operator,
        )
        db.add(log)
        await db.flush()

    @staticmethod
    async def list_batches(
        db: AsyncSession, page: int, page_size: int,
    ) -> tuple[list[ImportBatch], int]:
        stmt = select(ImportBatch).order_by(ImportBatch.created_at.desc())
        count_stmt = select(func.count()).select_from(stmt.subquery())
        total = await db.scalar(count_stmt) or 0
        offset = (page - 1) * page_size
        stmt = stmt.offset(offset).limit(page_size)
        result = await db.execute(stmt)
        batches = list(result.scalars().all())
        return batches, total

    @staticmethod
    async def get_batch(db: AsyncSession, batch_id: int) -> Optional[ImportBatch]:
        return await db.get(ImportBatch, batch_id)

    @staticmethod
    async def generate_error_report(db: AsyncSession, batch_id: int) -> io.BytesIO:
        batch = await db.get(ImportBatch, batch_id)
        if not batch:
            raise ValueError("批次不存在")

        stmt = select(ImportBatchRow).where(
            ImportBatchRow.batch_id == batch_id,
            ImportBatchRow.status == "error",
        ).order_by(ImportBatchRow.row_index)
        result = await db.execute(stmt)
        error_rows = list(result.scalars().all())

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "导入错误报告"

        raw_keys = []
        if error_rows and error_rows[0].raw_data:
            raw_keys = list(error_rows[0].raw_data.keys())
        headers = raw_keys + ["错误信息"]
        ImportBatchService._apply_header_style(ws, headers)

        for i, row_record in enumerate(error_rows, start=2):
            raw = row_record.raw_data or {}
            for col_idx, key in enumerate(raw_keys, start=1):
                ws.cell(row=i, column=col_idx, value=raw.get(key))
            ws.cell(row=i, column=len(headers), value=row_record.error_message)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
