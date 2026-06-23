"""商品列表物流数据工作台测试

覆盖：
- logistics_status 计算
- missing_logistics_fields 中文标签
- package_volume_weight_kg 计算
- cargo_type 筛选
- logistics_status 筛选 (complete/incomplete)
- Excel 导出/导入包含物流列
"""

import io
import openpyxl


async def _create_product(async_client, **overrides):
    payload = {
        "name": "物流列表测试商品",
        "unit": "件",
        "status": 0,
    }
    payload.update(overrides)
    resp = await async_client.post("/api/products", json=payload)
    assert resp.status_code == 200
    data = resp.json()
    assert data["code"] == 200
    return data["data"]


class TestProductLogisticsReadiness:
    async def test_product_list_returns_logistics_readiness(self, async_client):
        """完整物流信息的商品返回 complete"""
        product = await _create_product(
            async_client,
            name="完整物流商品",
            package_length_cm=25.0,
            package_width_cm=15.0,
            package_height_cm=10.0,
            package_weight_kg=0.8,
        )

        resp = await async_client.get(
            "/api/products", params={"page": 1, "page_size": 50}
        )
        assert resp.status_code == 200
        data = resp.json()
        assert data["code"] == 200
        records = {item["id"]: item for item in data["records"]}

        p = records[product["id"]]
        assert p["logistics_status"] == "complete"
        assert p["logistics_status_name"] == "物流完整"
        assert p["missing_logistics_fields"] == []
        assert p["package_volume_weight_kg"] == round(25.0 * 15.0 * 10.0 / 6000, 3)

    async def test_product_list_reports_missing_logistics_fields(self, async_client):
        """包装重量缺失时正确报告缺失字段，体积重仍然计算（仅依赖尺寸）"""
        product = await _create_product(
            async_client,
            name="缺重量商品",
            package_length_cm=20.0,
            package_width_cm=10.0,
            package_height_cm=5.0,
        )

        resp = await async_client.get(
            "/api/products", params={"page": 1, "page_size": 50}
        )
        assert resp.status_code == 200
        data = resp.json()
        assert data["code"] == 200
        records = {item["id"]: item for item in data["records"]}

        p = records[product["id"]]
        assert p["logistics_status"] == "incomplete"
        assert p["logistics_status_name"] == "物流不完整"
        assert "包装重量" in p["missing_logistics_fields"]
        # 有包装尺寸时可以计算体积重
        assert p["package_volume_weight_kg"] is not None

    async def test_product_list_reports_missing_all_package_fields(self, async_client):
        """没有任何包装信息时返回全部4个缺失字段"""
        product = await _create_product(
            async_client,
            name="全缺商品",
        )

        resp = await async_client.get(
            "/api/products", params={"page": 1, "page_size": 50}
        )
        assert resp.status_code == 200
        data = resp.json()
        assert data["code"] == 200
        records = {item["id"]: item for item in data["records"]}

        p = records[product["id"]]
        assert p["logistics_status"] == "incomplete"
        missing = set(p["missing_logistics_fields"])
        assert "包装长" in missing
        assert "包装宽" in missing
        assert "包装高" in missing
        assert "包装重量" in missing
        assert p["package_volume_weight_kg"] is None

    async def test_package_volume_weight_returns_none_when_missing_dimensions(
        self, async_client
    ):
        """包装尺寸任一缺失时体积重返回 None"""
        product = await _create_product(
            async_client,
            name="缺尺寸商品",
            package_length_cm=20.0,
            package_weight_kg=0.5,
        )

        resp = await async_client.get(f"/api/products/{product['id']}")
        assert resp.status_code == 200
        data = resp.json()["data"]
        assert data["package_volume_weight_kg"] is None

    async def test_package_volume_weight_with_zero_values(self, async_client):
        """部分包装尺寸缺失时视为不完整"""
        product = await _create_product(
            async_client,
            name="缺尺寸商品2",
            package_width_cm=10.0,
            package_height_cm=5.0,
            package_weight_kg=0.5,
        )

        resp = await async_client.get(f"/api/products/{product['id']}")
        assert resp.status_code == 200
        data = resp.json()["data"]
        assert data["logistics_status"] == "incomplete"
        assert "包装长" in data["missing_logistics_fields"]
        assert data["package_volume_weight_kg"] is None


class TestProductListFilters:
    async def test_product_list_filters_by_cargo_type(self, async_client):
        """按货品类型筛选"""
        await _create_product(async_client, name="普通商品", cargo_type="normal")
        await _create_product(async_client, name="电池商品", cargo_type="battery")
        await _create_product(async_client, name="液体商品", cargo_type="liquid")

        resp = await async_client.get(
            "/api/products",
            params={"cargo_type": "battery", "page": 1, "page_size": 50},
        )
        assert resp.status_code == 200
        data = resp.json()
        assert data["code"] == 200
        for item in data["records"]:
            assert item["cargo_type"] == "battery", (
                f"Expected battery, got {item['cargo_type']}"
            )

    async def test_product_list_filters_by_incomplete_logistics(self, async_client):
        """按 logistics_status=incomplete 筛选"""
        await _create_product(
            async_client,
            name="完整商品_筛选",
            package_length_cm=10.0,
            package_width_cm=10.0,
            package_height_cm=10.0,
            package_weight_kg=0.5,
        )
        incomplete = await _create_product(
            async_client,
            name="不完整商品_筛选",
            package_length_cm=20.0,
            package_width_cm=10.0,
            package_height_cm=5.0,
        )

        resp = await async_client.get(
            "/api/products",
            params={"logistics_status": "incomplete", "page": 1, "page_size": 50},
        )
        assert resp.status_code == 200
        data = resp.json()
        assert data["code"] == 200
        ids = [item["id"] for item in data["records"]]
        assert incomplete["id"] in ids
        for item in data["records"]:
            assert item["logistics_status"] == "incomplete"

    async def test_product_list_filters_by_complete_logistics(self, async_client):
        """按 logistics_status=complete 筛选"""
        complete = await _create_product(
            async_client,
            name="完整商品_B",
            package_length_cm=10.0,
            package_width_cm=10.0,
            package_height_cm=10.0,
            package_weight_kg=0.5,
        )
        await _create_product(
            async_client,
            name="不完整商品_B",
            package_length_cm=20.0,
            package_width_cm=10.0,
            package_height_cm=5.0,
        )

        resp = await async_client.get(
            "/api/products",
            params={"logistics_status": "complete", "page": 1, "page_size": 50},
        )
        assert resp.status_code == 200
        data = resp.json()
        assert data["code"] == 200
        ids = [item["id"] for item in data["records"]]
        assert complete["id"] in ids
        for item in data["records"]:
            assert item["logistics_status"] == "complete"


class TestProductExcelLogistics:
    async def _export_and_parse(self, async_client, params: dict = None) -> list[dict]:
        """导出 Excel 并解析为 dict 列表"""
        if params is None:
            params = {}
        resp = await async_client.get("/api/products/export", params=params)
        assert resp.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append({headers[i]: row[i] for i in range(len(headers))})
        return rows, headers

    async def test_export_includes_logistics_columns(self, async_client):
        """导出文件包含物流相关列"""
        await _create_product(
            async_client,
            name="导出测试商品",
            product_length_cm=10.0,
            product_width_cm=5.0,
            product_height_cm=3.0,
            product_weight_kg=0.2,
            package_length_cm=12.0,
            package_width_cm=6.0,
            package_height_cm=4.0,
            package_weight_kg=0.3,
            cargo_type="battery",
        )

        rows, headers = await self._export_and_parse(async_client)

        required_columns = [
            "商品长(cm)",
            "商品宽(cm)",
            "商品高(cm)",
            "商品重量(kg)",
            "包装长(cm)",
            "包装宽(cm)",
            "包装高(cm)",
            "包装重量(kg)",
            "货品类型",
            "物流状态",
        ]
        for col in required_columns:
            assert col in headers, f"导出缺少列: {col}"

        # 验证值正确
        assert len(rows) >= 1
        product_row = rows[0]
        assert product_row["商品长(cm)"] == 10.0
        assert product_row["包装长(cm)"] == 12.0
        assert product_row["货品类型"] == "带电"

    async def test_template_includes_logistics_columns(self, async_client):
        """导入模板包含所有物流列"""
        resp = await async_client.get("/api/products/export-template")
        assert resp.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]

        required_columns = [
            "商品名称",
            "副标题",
            "单位",
            "状态",
            "商品长(cm)",
            "商品宽(cm)",
            "商品高(cm)",
            "商品重量(kg)",
            "包装长(cm)",
            "包装宽(cm)",
            "包装高(cm)",
            "包装重量(kg)",
            "货品类型",
        ]
        for col in required_columns:
            assert col in headers, f"模板缺少列: {col}"

    async def test_import_reads_logistics_columns(self, async_client):
        """导入按表头名称正确读取物流字段"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "商品导入模板"
        headers = [
            "商品名称",
            "副标题",
            "单位",
            "状态",
            "商品长(cm)",
            "商品宽(cm)",
            "商品高(cm)",
            "商品重量(kg)",
            "包装长(cm)",
            "包装宽(cm)",
            "包装高(cm)",
            "包装重量(kg)",
            "货品类型",
        ]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)

        ws.cell(row=2, column=1, value="导入物流测试商品")
        ws.cell(row=2, column=3, value="件")
        ws.cell(row=2, column=4, value="上架")
        ws.cell(row=2, column=5, value=15.0)  # 商品长
        ws.cell(row=2, column=6, value=8.0)  # 商品宽
        ws.cell(row=2, column=7, value=5.0)  # 商品高
        ws.cell(row=2, column=8, value=0.3)  # 商品重量
        ws.cell(row=2, column=9, value=18.0)  # 包装长
        ws.cell(row=2, column=10, value=10.0)  # 包装宽
        ws.cell(row=2, column=11, value=6.0)  # 包装高
        ws.cell(row=2, column=12, value=0.5)  # 包装重量
        ws.cell(row=2, column=13, value="battery")  # 货品类型

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        resp = await async_client.post(
            "/api/products/import",
            files={
                "file": (
                    "test.xlsx",
                    output,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert resp.status_code == 200
        data = resp.json()
        assert data["code"] == 200
        assert data["data"]["imported"] == 1
        assert len(data["data"]["errors"]) == 0

        # 验证导入字段
        list_resp = await async_client.get(
            "/api/products",
            params={"name": "导入物流测试商品", "page": 1, "page_size": 10},
        )
        assert list_resp.status_code == 200
        list_data = list_resp.json()
        assert list_data["code"] == 200
        assert len(list_data["records"]) >= 1
        p = list_data["records"][0]
        assert p["product_length_cm"] == 15.0
        assert p["package_length_cm"] == 18.0
        assert p["cargo_type"] == "battery"

    async def test_import_reports_invalid_numeric_logistics(self, async_client):
        """导入时无效数值物流字段报行级错误"""
        wb = openpyxl.Workbook()
        ws = wb.active
        headers = [
            "商品名称",
            "副标题",
            "单位",
            "状态",
            "商品长(cm)",
            "商品宽(cm)",
            "商品高(cm)",
            "商品重量(kg)",
            "包装长(cm)",
            "包装宽(cm)",
            "包装高(cm)",
            "包装重量(kg)",
            "货品类型",
        ]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)

        ws.cell(row=2, column=1, value="无效数值导入商品")
        ws.cell(row=2, column=3, value="件")
        ws.cell(row=2, column=4, value="上架")
        ws.cell(row=2, column=9, value="abc")

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        resp = await async_client.post(
            "/api/products/import",
            files={
                "file": (
                    "bad.xlsx",
                    output,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert resp.status_code == 200
        data = resp.json()
        assert data["data"]["imported"] == 0
        assert len(data["data"]["errors"]) == 1
        assert (
            "包装长" in data["data"]["errors"][0] or "数字" in data["data"]["errors"][0]
        )
