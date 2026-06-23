"""导入批次 API 测试"""

import io
from uuid import uuid4

import openpyxl
import pytest
from httpx import AsyncClient


TEMPLATE_HEADERS = {
    "product": [
        "商品名称",
        "副标题",
        "单位",
        "状态",
        "商品长(cm)",
        "商品宽(cm)",
        "商品高(cm)",
        "商品重量(kg)",
        "包装长(cm)",
        "包装宽(cm)",
        "包装高(cm)",
        "包装重量(kg)",
        "货品类型",
    ],
    "sku": [
        "SKU编码",
        "所属商品ID",
        "条码",
        "规格描述",
        "规格值(JSON)",
        "销售价",
        "成本价",
        "市场价",
        "库存",
        "状态",
        "重量(kg)",
    ],
    "price": [
        "SKU编码",
        "价格类型",
        "价格",
        "生效时间",
        "失效时间",
    ],
    "inventory": [
        "SKU编码",
        "仓库",
        "货位",
        "数量",
        "模式",
        "安全库存",
    ],
}


def _make_excel(headers: list[str], rows: list[list]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    for row_idx, row in enumerate(rows, 2):
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


async def _preview_excel(
    async_client: AsyncClient,
    import_type: str,
    headers: list[str],
    rows: list[list],
    filename: str = "test.xlsx",
) -> dict:
    excel_bytes = _make_excel(headers, rows)
    resp = await async_client.post(
        "/api/import/preview",
        params={"type": import_type},
        files={
            "file": (
                filename,
                excel_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert resp.status_code == 200, f"Preview failed: {resp.text}"
    return resp.json()["data"]


async def _create_product_sku(async_client: AsyncClient, uid: str) -> tuple[int, str]:
    resp = await async_client.post("/api/products", json={"name": f"Test_{uid}"})
    assert resp.status_code == 200, f"Create product failed: {resp.text}"
    pid = resp.json()["data"]["id"]

    await async_client.post(
        f"/api/products/{pid}/specs",
        json={"specs": [{"name": "颜色", "values": ["标准"]}]},
    )

    resp = await async_client.post(f"/api/products/{pid}/skus/generate")
    assert resp.status_code == 200, f"Generate SKU failed: {resp.text}"
    sku_code = resp.json()["data"]["skus"][0]["code"]

    return pid, sku_code


# ========== Template Downloads ==========


@pytest.mark.asyncio
async def test_download_product_template(async_client: AsyncClient):
    resp = await async_client.get("/api/import/templates/product")
    assert resp.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    assert [cell.value for cell in ws[1]] == TEMPLATE_HEADERS["product"]


@pytest.mark.asyncio
async def test_download_sku_template(async_client: AsyncClient):
    resp = await async_client.get("/api/import/templates/sku")
    assert resp.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    assert [cell.value for cell in ws[1]] == TEMPLATE_HEADERS["sku"]


@pytest.mark.asyncio
async def test_download_price_template(async_client: AsyncClient):
    resp = await async_client.get("/api/import/templates/price")
    assert resp.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    assert [cell.value for cell in ws[1]] == TEMPLATE_HEADERS["price"]


@pytest.mark.asyncio
async def test_download_inventory_template(async_client: AsyncClient):
    resp = await async_client.get("/api/import/templates/inventory")
    assert resp.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    assert [cell.value for cell in ws[1]] == TEMPLATE_HEADERS["inventory"]


@pytest.mark.asyncio
async def test_download_template_invalid_type(async_client: AsyncClient):
    resp = await async_client.get("/api/import/templates/invalid")
    assert resp.status_code == 200
    assert resp.json()["code"] != 200


# ========== Preview ==========


@pytest.mark.asyncio
async def test_preview_valid_product(async_client: AsyncClient):
    uid = uuid4().hex[:6]
    data = await _preview_excel(
        async_client,
        "product",
        TEMPLATE_HEADERS["product"],
        [
            [
                f"产品{uid}",
                "副标题",
                "件",
                "上架",
                30,
                20,
                10,
                0.5,
                35,
                25,
                15,
                0.8,
                "normal",
            ],
        ],
    )
    assert data["type"] == "product"
    assert data["total_rows"] == 1
    assert data["valid_rows"] == 1
    assert data["error_rows"] == 0
    assert data["errors"] == []


@pytest.mark.asyncio
async def test_preview_empty_file(async_client: AsyncClient):
    excel_bytes = _make_excel(TEMPLATE_HEADERS["product"], [])
    resp = await async_client.post(
        "/api/import/preview",
        params={"type": "product"},
        files={
            "file": (
                "empty.xlsx",
                excel_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert resp.status_code == 200
    data = resp.json()["data"]
    assert data["total_rows"] == 0
    assert any("Excel文件为空" in e["error_message"] for e in data["errors"])


@pytest.mark.asyncio
async def test_preview_with_validation_errors(async_client: AsyncClient):
    data = await _preview_excel(
        async_client,
        "product",
        TEMPLATE_HEADERS["product"],
        [
            [
                "",
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
            ],
        ],
    )
    assert data["total_rows"] == 1
    assert data["valid_rows"] == 0
    assert data["error_rows"] == 1
    assert any("商品名称不能为空" in e["error_message"] for e in data["errors"])


@pytest.mark.asyncio
async def test_preview_inventory_invalid_mode(async_client: AsyncClient):
    data = await _preview_excel(
        async_client,
        "inventory",
        TEMPLATE_HEADERS["inventory"],
        [
            ["SKU001", "主仓", "A-1", 100, "bad_mode", 10],
        ],
    )
    assert data["total_rows"] == 1
    assert data["error_rows"] == 1
    assert any("模式无效" in e["error_message"] for e in data["errors"])


# ========== Commit ==========


@pytest.mark.asyncio
async def test_preview_then_commit_product(async_client: AsyncClient):
    uid = uuid4().hex[:6]

    data = await _preview_excel(
        async_client,
        "product",
        TEMPLATE_HEADERS["product"],
        [
            [
                f"产品{uid}",
                None,
                "件",
                "上架",
                30,
                20,
                10,
                0.5,
                None,
                None,
                None,
                None,
                "normal",
            ],
            [
                f"产品{uid}_B",
                None,
                "件",
                "草稿",
                10,
                10,
                10,
                0.2,
                None,
                None,
                None,
                None,
                "normal",
            ],
        ],
    )
    batch_id = data["batch_id"]

    commit_resp = await async_client.post(f"/api/import/commit/{batch_id}")
    assert commit_resp.status_code == 200
    result = commit_resp.json()["data"]
    assert result["success_count"] == 2
    assert result["error_count"] == 0

    detail_resp = await async_client.get(f"/api/import/batches/{batch_id}")
    assert detail_resp.status_code == 200
    assert detail_resp.json()["data"]["status"] == "committed"

    list_resp = await async_client.get("/api/products", params={"name": uid})
    assert list_resp.status_code == 200
    product_names = [p["name"] for p in list_resp.json()["records"]]
    assert any(f"产品{uid}" in n for n in product_names)


@pytest.mark.asyncio
async def test_preview_then_commit_price(async_client: AsyncClient):
    uid = uuid4().hex[:6]
    pid, sku_code = await _create_product_sku(async_client, uid)

    data = await _preview_excel(
        async_client,
        "price",
        TEMPLATE_HEADERS["price"],
        [
            [sku_code, "sale_price", 99.99, None, None],
            [sku_code, "cost_price", 50.00, None, None],
        ],
    )
    batch_id = data["batch_id"]

    commit_resp = await async_client.post(f"/api/import/commit/{batch_id}")
    assert commit_resp.status_code == 200
    result = commit_resp.json()["data"]
    assert result["success_count"] == 2
    assert result["error_count"] == 0

    skus_resp = await async_client.get(f"/api/products/{pid}/skus")
    assert skus_resp.status_code == 200
    sku_id = skus_resp.json()["data"][0]["id"]

    prices_resp = await async_client.get(f"/api/skus/{sku_id}/prices")
    assert prices_resp.status_code == 200
    prices = prices_resp.json()["data"]
    price_map = {p["price_type"]: p["price"] for p in prices}
    assert float(price_map["sale_price"]) == 99.99
    assert float(price_map["cost_price"]) == 50.00


@pytest.mark.asyncio
async def test_commit_without_preview(async_client: AsyncClient):
    resp = await async_client.post("/api/import/commit/999999")
    assert resp.status_code == 200
    assert resp.json()["code"] != 200


@pytest.mark.asyncio
async def test_double_commit(async_client: AsyncClient):
    uid = uuid4().hex[:6]

    data = await _preview_excel(
        async_client,
        "product",
        TEMPLATE_HEADERS["product"],
        [
            [
                f"产品{uid}",
                None,
                "件",
                "上架",
                30,
                20,
                10,
                0.5,
                None,
                None,
                None,
                None,
                "normal",
            ],
        ],
    )
    batch_id = data["batch_id"]

    resp1 = await async_client.post(f"/api/import/commit/{batch_id}")
    assert resp1.status_code == 200

    resp2 = await async_client.post(f"/api/import/commit/{batch_id}")
    assert resp2.status_code == 200
    assert resp2.json()["code"] != 200


# ========== Batch Listing ==========


@pytest.mark.asyncio
async def test_list_batches(async_client: AsyncClient):
    uid = uuid4().hex[:6]
    await _preview_excel(
        async_client,
        "product",
        TEMPLATE_HEADERS["product"],
        [
            [
                f"产品{uid}",
                None,
                "件",
                "上架",
                30,
                20,
                10,
                0.5,
                None,
                None,
                None,
                None,
                "normal",
            ],
        ],
    )

    resp = await async_client.get("/api/import/batches")
    assert resp.status_code == 200
    result = resp.json()
    assert len(result["records"]) >= 1
    assert result["total"] >= 1

    batch_types = {b["type"] for b in result["records"]}
    assert "product" in batch_types


@pytest.mark.asyncio
async def test_get_batch_detail(async_client: AsyncClient):
    uid = uuid4().hex[:6]
    data = await _preview_excel(
        async_client,
        "product",
        TEMPLATE_HEADERS["product"],
        [
            [
                f"产品{uid}",
                None,
                "件",
                "上架",
                30,
                20,
                10,
                0.5,
                None,
                None,
                None,
                None,
                "normal",
            ],
        ],
    )
    batch_id = data["batch_id"]

    resp = await async_client.get(f"/api/import/batches/{batch_id}")
    assert resp.status_code == 200
    detail = resp.json()["data"]
    assert detail["id"] == batch_id
    assert detail["type"] == "product"
    assert detail["status"] == "previewed"
    assert detail["total_rows"] == 1


@pytest.mark.asyncio
async def test_download_error_report(async_client: AsyncClient):
    data = await _preview_excel(
        async_client,
        "product",
        TEMPLATE_HEADERS["product"],
        [
            [
                "",
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
            ],
        ],
    )
    batch_id = data["batch_id"]

    resp = await async_client.get(f"/api/import/batches/{batch_id}/errors")
    assert resp.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    assert ws.title == "导入错误报告"
    header_cells = [cell.value for cell in ws[1]]
    assert "错误信息" in header_cells
    assert ws.max_row >= 2
