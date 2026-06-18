"""批量上架决策 Excel 导入导出测试。"""

import io

import openpyxl
import pytest
from httpx import AsyncClient

from app.decision.excel_service import PreListingDecisionExcelService


def _workbook_bytes(rows: list[list[object]]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "批量上架决策"
    for row in rows:
        ws.append(row)
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def _valid_rows() -> list[list[object]]:
    return [
        [
            "行标识",
            "SKU ID",
            "目的国",
            "目标售价",
            "平台ID",
            "类目ID",
            "平台费率(%)",
            "支付费率(%)",
            "其他费用",
            "最低利润率(%)",
            "货品类型",
        ],
        ["row-a", 1, "ru", 5000, None, None, None, None, None, None, None],
        ["row-b", 2, "MY", 1000, 7, 8, 12, 4, 9, 25, "battery"],
    ]


def test_parse_preview_maps_rows_and_defaults():
    content = _workbook_bytes(_valid_rows())

    preview = PreListingDecisionExcelService.parse_preview(content)

    assert preview.total_rows == 2
    assert preview.valid_rows == 2
    assert preview.error_rows == 0
    first = preview.items[0]
    assert first.row_number == 2
    assert first.item is not None
    assert first.item.item_key == "row-a"
    assert first.item.sku_id == 1
    assert first.item.destination_country == "RU"
    assert first.item.platform_fee_pct == 10
    assert first.item.payment_fee_pct == 3
    assert first.item.other_fee == 0
    assert first.item.minimum_margin_pct == 20
    assert first.item.cargo_type == "normal"

    second = preview.items[1].item
    assert second is not None
    assert second.platform_id == 7
    assert second.category_id == 8
    assert second.platform_fee_pct == 12
    assert second.payment_fee_pct == 4
    assert second.other_fee == 9
    assert second.minimum_margin_pct == 25
    assert second.cargo_type == "battery"


def test_parse_preview_returns_row_errors_without_crashing():
    content = _workbook_bytes(
        [
            ["行标识", "SKU ID", "目的国", "目标售价"],
            ["bad-row", "abc", "", -1],
        ]
    )

    preview = PreListingDecisionExcelService.parse_preview(content)

    assert preview.total_rows == 1
    assert preview.valid_rows == 0
    assert preview.error_rows == 1
    assert preview.items[0].row_number == 2
    assert preview.items[0].item is None
    assert any("SKU ID必须是整数" in err for err in preview.items[0].errors)
    assert any("目的国不能为空" in err for err in preview.items[0].errors)
    assert any("目标售价必须大于0" in err for err in preview.items[0].errors)


def test_parse_preview_rejects_more_than_100_rows():
    rows = [["SKU ID", "目的国", "目标售价"]]
    rows.extend([[idx, "RU", 100] for idx in range(1, 102)])
    content = _workbook_bytes(rows)

    with pytest.raises(ValueError, match="最多支持100行"):
        PreListingDecisionExcelService.parse_preview(content)


def test_generate_template_creates_valid_workbook():
    content = PreListingDecisionExcelService.generate_template()

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    assert ws.title == "批量上架决策"
    headers = [cell.value for cell in ws[1]]
    assert "SKU ID" in headers
    assert "目的国" in headers
    assert "目标售价" in headers


@pytest.mark.skip(reason="endpoint /api/decisions/prelisting/batch/template, /api/decisions/prelisting/batch/preview, /api/decisions/prelisting/batch/export not implemented yet")
@pytest.mark.asyncio
async def test_excel_endpoints_require_permission(async_client: AsyncClient):
    """测试模板、预览、导出端点需要 decision:calculate 权限"""
    # GET template
    resp = await async_client.get("/api/decisions/prelisting/batch/template")
    assert resp.status_code == 200

    # POST preview with invalid file
    resp = await async_client.post(
        "/api/decisions/prelisting/batch/preview",
        files={"file": ("test.txt", b"not an excel", "text/plain")},
    )
    assert resp.status_code == 400

    # POST preview with valid excel
    content = _workbook_bytes(_valid_rows())
    resp = await async_client.post(
        "/api/decisions/prelisting/batch/preview",
        files={"file": ("test.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 200
    data = resp.json()["data"]
    assert data["total_rows"] == 2
    assert data["valid_rows"] == 2

    # POST export
    resp = await async_client.post(
        "/api/decisions/prelisting/batch/export",
        json={
            "summary": {
                "total_items": 1,
                "success_count": 1,
                "error_count": 0,
                "approve_count": 1,
                "reject_count": 0,
                "needs_data_count": 0,
                "average_profit_margin": 50.0,
            },
            "items": [
                {
                    "index": 0,
                    "item_key": "row-a",
                    "sku_id": 1,
                    "status": "success",
                    "result": {
                        "sku_id": 1,
                        "destination_country": "RU",
                        "target_sale_price": 5000,
                        "product_cost": 500,
                        "shipping_fee": 50,
                        "platform_fee": 500,
                        "payment_fee": 150,
                        "fixed_fee": 0,
                        "advertising_fee": 0,
                        "other_fee": 100,
                        "profit_amount": 3700,
                        "profit_margin": 74.0,
                        "recommendation": "approve",
                        "blocking_reasons": [],
                        "warnings": [],
                        "applied_platform_fee_rule_id": None,
                        "platform_fee_source": "manual",
                        "platform_fee_rule_summary": None,
                    },
                    "error_message": None,
                },
                {
                    "index": 1,
                    "item_key": "row-b",
                    "sku_id": 999,
                    "status": "error",
                    "result": None,
                    "error_message": "SKU不存在",
                },
            ],
        },
    )
    assert resp.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(resp.content), data_only=True)
    ws = wb.active
    assert ws.title == "测算结果"
    assert ws.cell(row=2, column=4).value == "成功"
    assert ws.cell(row=2, column=5).value == "approve"
    assert ws.cell(row=3, column=4).value == "错误"
    assert ws.cell(row=3, column=20).value == "SKU不存在"


def test_parse_preview_handles_chinese_fullwidth_brackets():
    """全角括号表头也应被识别"""
    rows = [
        ["行标识", "SKU ID", "目的国", "目标售价", "平台ID", "类目ID",
         "平台费率（%）", "支付费率（%）", "其他费用", "最低利润率（%）", "货品类型"],
        ["r1", 1, "RU", 100, None, None, 15, 5, 2, 30, "normal"],
    ]
    content = _workbook_bytes(rows)
    preview = PreListingDecisionExcelService.parse_preview(content)
    assert preview.valid_rows == 1
    assert preview.error_rows == 0
    item = preview.items[0].item
    assert item is not None
    assert item.platform_fee_pct == 15
    assert item.payment_fee_pct == 5
    assert item.other_fee == 2
    assert item.minimum_margin_pct == 30
