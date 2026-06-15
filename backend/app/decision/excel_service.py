"""批量上架决策 Excel 服务。"""

import io
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.decision.schemas import (
    PreListingDecisionBatchItem,
    PreListingDecisionBatchResponse,
    PreListingDecisionExcelPreviewResponse,
    PreListingDecisionExcelPreviewRow,
)


def _text(value: Any) -> str:
    return str(value).strip() if value not in (None, "") else ""


def _normalize_header(value: Any) -> str:
    return _text(value).replace("（", "(").replace("）", ")")


def _int_or_none(value: Any, field_name: str, errors: list[str]) -> int | None:
    if value in (None, ""):
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        errors.append(f"{field_name}必须是整数")
        return None


def _required_int(value: Any, field_name: str, errors: list[str]) -> int | None:
    parsed = _int_or_none(value, field_name, errors)
    if parsed is None and value in (None, ""):
        errors.append(f"{field_name}不能为空")
    return parsed


def _float_or_default(value: Any, field_name: str, default: float, errors: list[str]) -> float:
    if value in (None, ""):
        return default
    try:
        parsed = float(value)
    except (TypeError, ValueError):
        errors.append(f"{field_name}必须是数字")
        return default
    if parsed < 0:
        errors.append(f"{field_name}不能小于0")
    return parsed


def _positive_float(value: Any, field_name: str, errors: list[str]) -> float | None:
    if value in (None, ""):
        errors.append(f"{field_name}不能为空")
        return None
    try:
        parsed = float(value)
    except (TypeError, ValueError):
        errors.append(f"{field_name}必须是数字")
        return None
    if parsed <= 0:
        errors.append(f"{field_name}必须大于0")
    return parsed


class PreListingDecisionExcelService:
    """批量上架决策 Excel 模板、预览和结果导出。"""

    SHEET_NAME = "批量上架决策"
    RESULT_SHEET_NAME = "测算结果"
    MAX_ROWS = 100
    HEADERS = [
        "行标识",
        "SKU ID",
        "目的国",
        "目标售价",
        "平台ID",
        "类目ID",
        "平台费率(%)",
        "支付费率(%)",
        "其他费用",
        "最低利润率(%)",
        "货品类型",
    ]
    RESULT_HEADERS = [
        "行号",
        "行标识",
        "SKU ID",
        "状态",
        "建议",
        "目的国",
        "目标售价",
        "商品成本",
        "运费",
        "平台费",
        "支付费",
        "固定交易费",
        "广告预留",
        "其他费用",
        "利润金额",
        "利润率(%)",
        "费用来源",
        "阻断原因",
        "警告",
        "错误",
    ]
    HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    @staticmethod
    def _style_header(ws, headers: list[str]) -> None:
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = PreListingDecisionExcelService.HEADER_FILL
            cell.font = PreListingDecisionExcelService.HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.border = PreListingDecisionExcelService.THIN_BORDER

    @staticmethod
    def _set_col_widths(ws, widths: list[int]) -> None:
        for col, width in enumerate(widths, 1):
            col_letter = openpyxl.utils.get_column_letter(col)
            ws.column_dimensions[col_letter].width = width

    # ── Template ────────────────────────────────────────────────────────────

    @staticmethod
    def generate_template() -> bytes:
        """生成空白模板工作簿，返回字节。"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = PreListingDecisionExcelService.SHEET_NAME

        PreListingDecisionExcelService._style_header(ws, PreListingDecisionExcelService.HEADERS)
        PreListingDecisionExcelService._set_col_widths(ws, [12, 10, 10, 12, 10, 10, 12, 12, 10, 14, 12])

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    # ── Preview ─────────────────────────────────────────────────────────────

    @staticmethod
    def parse_preview(content: bytes) -> PreListingDecisionExcelPreviewResponse:
        """解析上传的 Excel 文件，返回预览响应（含行级错误）。"""
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet()

        # 读取所有行
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            return PreListingDecisionExcelPreviewResponse(
                total_rows=0, valid_rows=0, error_rows=0, items=[]
            )

        # 构建表头映射（支持全角括号）
        raw_headers = [str(c.value) if c.value is not None else "" for c in ws[1]]
        header_map: dict[str, int] = {}
        for idx, h in enumerate(raw_headers):
            normalized = _normalize_header(h)
            if normalized:
                header_map[normalized] = idx

        # 解析数据行
        items: list[PreListingDecisionExcelPreviewRow] = []
        data_rows = all_rows[1:]  # 跳过表头
        non_empty_count = 0

        for row_idx, row_values in enumerate(data_rows, start=2):
            # 跳过完全空行
            if all(v in (None, "") for v in row_values):
                continue

            non_empty_count += 1
            if non_empty_count > PreListingDecisionExcelService.MAX_ROWS:
                raise ValueError(f"最多支持{PreListingDecisionExcelService.MAX_ROWS}行数据")

            errors: list[str] = []
            parsed = PreListingDecisionExcelService._parse_row(row_values, header_map, errors)

            if errors:
                items.append(
                    PreListingDecisionExcelPreviewRow(
                        row_number=row_idx,
                        item=None,
                        errors=errors,
                    )
                )
            else:
                items.append(
                    PreListingDecisionExcelPreviewRow(
                        row_number=row_idx,
                        item=parsed,
                        errors=[],
                    )
                )

        valid_count = sum(1 for it in items if it.item is not None)
        error_count = sum(1 for it in items if it.errors)

        return PreListingDecisionExcelPreviewResponse(
            total_rows=len(items),
            valid_rows=valid_count,
            error_rows=error_count,
            items=items,
        )

    @staticmethod
    def _cell(row_values: tuple, header_map: dict, name: str) -> Any:
        idx = header_map.get(_normalize_header(name))
        if idx is None or idx >= len(row_values):
            return None
        return row_values[idx]

    @staticmethod
    def _parse_row(
        row_values: tuple,
        header_map: dict,
        errors: list[str],
    ) -> PreListingDecisionBatchItem | None:
        """将 Excel 行解析为 PreListingDecisionBatchItem。"""
        # 必填字段
        sku_id = _required_int(
            PreListingDecisionExcelService._cell(row_values, header_map, "SKU ID"),
            "SKU ID",
            errors,
        )
        destination_country = _text(
            PreListingDecisionExcelService._cell(row_values, header_map, "目的国")
        )
        if not destination_country:
            errors.append("目的国不能为空")

        target_sale_price = _positive_float(
            PreListingDecisionExcelService._cell(row_values, header_map, "目标售价"),
            "目标售价",
            errors,
        )

        if errors:
            return None

        # 可选字段
        item_key = _text(
            PreListingDecisionExcelService._cell(row_values, header_map, "行标识")
        ) or None

        platform_id = _int_or_none(
            PreListingDecisionExcelService._cell(row_values, header_map, "平台ID"),
            "平台ID",
            errors,
        )
        category_id = _int_or_none(
            PreListingDecisionExcelService._cell(row_values, header_map, "类目ID"),
            "类目ID",
            errors,
        )

        platform_fee_pct = _float_or_default(
            PreListingDecisionExcelService._cell(row_values, header_map, "平台费率(%)"),
            "平台费率(%)",
            10,
            errors,
        )
        payment_fee_pct = _float_or_default(
            PreListingDecisionExcelService._cell(row_values, header_map, "支付费率(%)"),
            "支付费率(%)",
            3,
            errors,
        )
        other_fee = _float_or_default(
            PreListingDecisionExcelService._cell(row_values, header_map, "其他费用"),
            "其他费用",
            0,
            errors,
        )
        minimum_margin_pct = _float_or_default(
            PreListingDecisionExcelService._cell(row_values, header_map, "最低利润率(%)"),
            "最低利润率(%)",
            20,
            errors,
        )

        raw_cargo = _text(
            PreListingDecisionExcelService._cell(row_values, header_map, "货品类型")
        )
        cargo_type = raw_cargo if raw_cargo else "normal"

        return PreListingDecisionBatchItem(
            item_key=item_key or None,
            sku_id=sku_id,
            destination_country=destination_country.upper(),
            target_sale_price=target_sale_price,
            platform_id=platform_id,
            category_id=category_id,
            platform_fee_pct=platform_fee_pct,
            payment_fee_pct=payment_fee_pct,
            other_fee=other_fee,
            minimum_margin_pct=minimum_margin_pct,
            cargo_type=cargo_type,
        )

    # ── Export ──────────────────────────────────────────────────────────────

    @staticmethod
    def export_results(data: PreListingDecisionBatchResponse) -> bytes:
        """将批量决策结果导出为 Excel 工作簿，返回字节。"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = PreListingDecisionExcelService.RESULT_SHEET_NAME

        PreListingDecisionExcelService._style_header(ws, PreListingDecisionExcelService.RESULT_HEADERS)
        PreListingDecisionExcelService._set_col_widths(ws, [8, 14, 10, 8, 12, 10, 12, 12, 10, 10, 10, 12, 12, 10, 12, 12, 16, 20, 30, 30])

        for row_idx, item in enumerate(data.items, start=2):
            status_label = "成功" if item.status == "success" else "错误"
            rec = item.result.recommendation if item.result else ""
            dest = item.result.destination_country if item.result else ""
            price = item.result.target_sale_price if item.result else ""
            cost = item.result.product_cost if item.result else ""
            shipping = item.result.shipping_fee if item.result else ""
            pfee = item.result.platform_fee if item.result else ""
            payfee = item.result.payment_fee if item.result else ""
            fixed = item.result.fixed_fee if item.result else ""
            ad = item.result.advertising_fee if item.result else ""
            other = item.result.other_fee if item.result else ""
            profit = item.result.profit_amount if item.result else ""
            margin = item.result.profit_margin if item.result else ""
            fee_source = item.result.platform_fee_source if item.result else ""
            reasons = "；".join(item.result.blocking_reasons) if item.result and item.result.blocking_reasons else ""
            warnings = "；".join(item.result.warnings) if item.result and item.result.warnings else ""
            err_msg = item.error_message or ""

            vals = [
                item.index + 1,
                item.item_key,
                item.sku_id,
                status_label,
                rec,
                dest,
                price,
                cost,
                shipping,
                pfee,
                payfee,
                fixed,
                ad,
                other,
                profit,
                margin,
                fee_source,
                reasons,
                warnings,
                err_msg,
            ]
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.border = PreListingDecisionExcelService.THIN_BORDER

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
